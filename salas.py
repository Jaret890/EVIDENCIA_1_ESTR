from datetime import datetime, timedelta
import random
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

#Listas
clientes = []
salas = []
reservaciones = []


#Clases
class Cliente:
    def __init__(self, clave, nombre, apellidos):
        self.clave = clave
        self.nombre = nombre
        self.apellidos = apellidos

class Sala:
    def __init__(self, clave, nombre, cupo):
        self.clave = clave
        self.nombre = nombre
        self.cupo = cupo

class Reservacion:
    def __init__(self, folio, fecha, turno, sala, cliente, evento):
        self.folio = folio
        self.fecha = fecha
        self.turno = turno
        self.sala = sala
        self.cliente = cliente
        self.evento = evento 
        
#Generador para las claves
def clave_cliente():
    while True:
        clave = f"C{random.randint(100,999)}"
        if not any(c.clave == clave for c in clientes):
            return clave

def clave_sala():
    while True:
        clave = f"S{random.randint(100,999)}"
        if not any(s.clave == clave for s in salas):
            return clave

def generar_folio():
    while True:
        folio = f"F{random.randint(10000,99999)}"
        if not any(r.folio == folio for r in reservaciones):
            return folio

#Constantes
TURNOS = {"1": "Matutino", "2": "Vespertino", "3": "Nocturno"}

FORMATO_FECHA = "%d/%m/%Y"

#Funciones
def registar_reservacion():
    if not clientes or not salas:
        print("--No hay ningun cliente/sala registrados--")
        return

    ordenados = sorted(clientes, key=lambda c: (c.apellidos.lower(), c.nombre.lower()))
    print("Clientes Registrados: ")
    print(f"{'Clave':6} | {'Apellidos':20} | {'Nombre':15}")
    print("-"*50)
    for c in ordenados:
        print(f"{c.clave:6} | {c.apellidos:20} | {c.nombre:15} ") 

    while True:
        clave_cliente = input("Clave del cliente ('0' para cancelar): ")
        if clave_cliente == "0":
            return
        cliente = next((c for c in clientes if c.clave == clave_cliente), None)
        if cliente:
            break
        print("--Clave no encontrada--")

    hoy = datetime.now().date()
    minimo = hoy + timedelta(days=2)
    while True:
        fecha_str = input(f"Fecha del evento ({FORMATO_FECHA}), minimo {minimo.strftime(FORMATO_FECHA)}: ")
        try:
            fecha = datetime.strptime(fecha_str, FORMATO_FECHA).date()
            if fecha < minimo:
                print("La reservacion debe de hacerse al menos dos dias antes")
            else:
                break
        except:
            print("Fecha invalida")

    print("--Salas con disponibilidad: ")
    disponibles = []
    for s in salas:
        libres = [t for t in TURNOS.values() if not any(r.sala.clave == s.clave and r.fecha == fecha and r.turno == t for r in reservaciones)]
        if libres:
            disponibles.append((s, libres))
            print(f"{s.clave} - {s.nombre} (cupo{s.cupo}) | Turnos: {', '.join(libres)}")
    
    if not disponibles:
        print("--No hay salas disponibles en esa fecha--")
        return
    
    clave_sala = input("Clave de sala ('0' para cancelar): ")
    if clave_sala == "0":
        return
    sala = next((s for s in salas if s.clave == clave_sala), None)
    if not sala:
        print("--Sala no encontrada--")
        return
    
    libres = [t for t in TURNOS.values() if not any(r.sala.clave == sala.clave and r.fecha == fecha and r.turno == t for r in reservaciones)]
    print("--Turnos disponibles--")
    for k, v in TURNOS.items():
        if v in libres:
            print(f"{k}. {v}")

    opcion_turno = input("-Elige el turno (1/2/3): ")
    if opcion_turno not in TURNOS or TURNOS[opcion_turno] not in libres:
        print("--Turno invalido--")
        return
    
    nombre_evento = input("-Nombre del evento: ")
    if not nombre_evento:
        print("--El nombre no puede estar vacio--")
        return
    turno = TURNOS[opcion_turno]
    
    folio = generar_folio()
    reservaciones.append(Reservacion(folio, fecha, turno, sala, cliente, nombre_evento))
    print(f"--Reservacion registrada con folio {folio}")

def editar_evento():
    if not reservaciones:
        print("--No hay reservaciones registradas--")
        return
    fecha_inicio_str = input(f"Fecha inicio ({FORMATO_FECHA}): ")
    fecha_fin_str = input(f"Fecha fin ({FORMATO_FECHA}): ")
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, FORMATO_FECHA).date()
        fecha_fin = datetime.strptime(fecha_fin_str, FORMATO_FECHA).date()
    except:
        print("--Fechas invalidas--")
        return
    if fecha_inicio > fecha_fin:
        print("--La fecha inicial no puede se mayor que la final--")
        return
    
    seleccionadas = [r for r in reservaciones if fecha_inicio <= r.fecha <= fecha_fin]
    if not seleccionadas:
        print("--No existen reservacines en ese rango--")
        return
    
    while True:
        print("--Reservaciones en rango:")
        for r in seleccionadas:
            print(f"Folio: {r.folio}, Evento: {r.evento}, Fecha: {r.fecha.strftime(FORMATO_FECHA)}")
        folio = input("Folio a editar ('0' para cancelar): ")
        if folio == "0":
            return
        res = next((r for r in seleccionadas if r.folio == folio), None)
        if res:
            nuevo_nombre = input("Nuevo nombre: ")
            if nuevo_nombre:
                res.nombre_evento = nuevo_nombre
                print("--Evento actualizado--")
                return
            else:
                print("--El nombre no puede estar vacio--")
        else:
            print("--Folio no valido--")

def consultar_por_fecha():
    if not reservaciones:
        print("--No hay reservaciones registradas--")
        return
    fecha_str = input(f"Fecha a consultar ({FORMATO_FECHA}): ")
    try:
        fecha = datetime.strptime(fecha_str, FORMATO_FECHA).date()
    except:
        print("--Fecha invalida--")
        return
    encontrados = [r for r in reservaciones if r.fecha == fecha]
    if not encontrados:
        print("--No hay reservaciones en esta fecha--")
        return
    print(f"--Reservaciones para {fecha.strftime(FORMATO_FECHA)}")
    print(f"{'Folio':8} | {'Evento':25} | {'Sala':15} | {'Turno':10} | {'Cliente':20}")
    print("-" * 85)
    for r in encontrados:
        print(f"Folio: {r.folio}, Evento: {r.evento}, Sala: {r.sala.nombre}, Turno: {r.turno}, Cliente: {r.cliente.nombre} {r.cliente.apellidos}")
    
    exportar = input("¿Desea exportar el reporte? (S/N): ").upper()
    if exportar != "S":
        return
    
    print("--Formatos disponibles--")
    print("1-CSV")
    print("2-JSON")
    print("3-Excel")
    formato = input("Selecciona el formato deseado (1/2/3): ")
    
    if formato == "1":
        pass
    elif formato == "2":
        pass
    elif formato == "3":
        pass
    else:
        print("--Opcion invalida--")


def exportar_csv(datos, fecha):
    nombre_archivo = f"reporte_{fecha.strftime('%Y%m%d')}.csv"
    with open(nombre_archivo, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["Folio", "Evento", "Sala", "Turno", "Cliente"])
        for r in datos:
            writer.writerow([r.folio, r.evento, r.sala.nombre, r.turno, f"{r.cliente.nombre} {r.cliente.apellidos}"])
    print(f"--Reporte exportado como CSV: {nombre_archivo}--")


def exportar_json(datos, fecha):
    nombre_archivo = f"reporte_{fecha.strftime('%Y%m%d')}.json"
    data = []
    for r in datos:
        data.append({"Folio": r.folio,
                     "Evento": r.evento,
                     "Sala": r.sala,
                     "Turno": r.turno,
                     "Cliente": f"{r.cliente.nombre} {r.cliente.apellidos}"
                     })
        
    with open(nombre_archivo, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
    print(f"--Reporte exportado como JSON: {nombre_archivo}--")


def exportar_excel(datos, fecha):
    nombre_archivo = f"reporte_{fecha.strfime('%Y%m%d')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservaciones"

    titulo = f"Reservaciones - {fecha.strftime(FORMATO_FECHA)}"
    ws.merge_cells("A1:E1")
    ws["A1"] = titulo
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    encabezados = ["Folio", "Evento", "Sala", "Turno", "Cliente"]
    ws.append(encabezados)

    bold_font = Font(bold=True)
    borde = Border(bottom=Side(style="thick"))
    centrado = Alignment(horizontal="center")


    for columna in range(1,len(encabezados) +1):
        celda = ws.cell(row=2, column=columna)
        celda.font = bold_font
        celda.border = borde
        celda.alignment = centrado
    
    for r in datos:
        ws.append([r.folio, r.evento, r.sala.nombre, r.turno, f"{r.cliente.nombre} {r.cliente.apellidos}"])

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for celda in row:
            celda.alignment = centrado
    
    for col in ws.columns:
        max_len = max(len(str(celda.value)) if celda.value else 0 for celda in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(nombre_archivo)
    print(f"--Reporte exportado como Excel: {nombre_archivo}--")


def registrar_cliente():
    clave =  clave_cliente()
    nombre = input("-Nombre del cliente: ")
    apellidos = input("-Apellidos: ")
    clientes.append(Cliente(clave, nombre, apellidos))
    print(f"--Cliente registrado con clave {clave}--")

def registrar_sala():
    clave = clave_sala()
    nombre = input("-Nombre de la sala: ")
    cupo = int(input("-Cupo maximo: "))
    salas.append(Sala(clave, nombre, cupo))
    print(f"--Sala registrada con clave {clave}--")
     

    

#Menu
while True:
    print("--Menu de Reservas--")
    print("1.-Registrar reservacion de sala")
    print("2.-Editar el nombre del evento de una reservacion ya hecha")
    print("3.-Consultar las reservaciones por una fecha en especifico")
    print("4.-Registrar un nuevo cliente")
    print("5.-Registrar un sala")
    print("6.-Salir")
    opcion = int(input("--Elige una opcion: "))

    if opcion == 1:
        registar_reservacion()
    elif opcion == 2:
        editar_evento()
    elif opcion == 3:
        consultar_por_fecha()
    elif opcion == 4:
        registrar_cliente()
    elif opcion == 5:
        registrar_sala()
    elif opcion == 6:
        print("--Adios--")
        break